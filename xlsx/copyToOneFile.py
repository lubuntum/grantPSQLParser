import array
import glob
import os.path
import xml.etree.ElementTree

from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

IGNORE_LIST = ['tags']


def get_first_sheet(wb: Workbook):
    return wb[wb.sheetnames[0]]


# Для создания первой строки в новом файле вызывается в самом начале, возвращает имена ячеек первой строки
def get_columns_names(sheet: Worksheet, ignore_list: []):
    column_names = []
    for cell in sheet[1]:
        if cell.value in ignore_list: continue
        column_names.append(cell.value)
    return column_names


# Отбирает из файла нужные столбцы и сохраняет их индекс для дальнейшей работы, вызывается много раз, возвращает индексы
def get_sheet_indexes(sheet: Worksheet, ignore_list: []):
    indexes = []
    for cell in sheet[1]:
        if cell.value in ignore_list: continue
        indexes.append(cell.column)
    return indexes


# Написать функцию для создания результирующего xlsx файла
# Написать функцию для запроса всех файлов из path
def get_xlsx_from_folder(path: str):
    files_path = glob.glob(f'{path}/*.xlsx')
    return files_path


def get_current_folder():
    script_path = os.path.abspath(__file__)
    return os.path.dirname(script_path)


def create_end_file(filename, column_names: []):
    end_file = Workbook()
    end_file_sheet = end_file.active
    end_file_sheet.append(column_names)
    end_file.save(filename)
    return end_file


if __name__ == '__main__':
    xlsx_docs = get_xlsx_from_folder(get_current_folder())
    print(xlsx_docs, get_current_folder())
    # По первому однотипному файлу из массива создаем конечный файл с той же первой строкой + учет игнор листа
    column_names = get_columns_names(get_first_sheet(load_workbook(xlsx_docs[0])), IGNORE_LIST)
    result = create_end_file('Result.xlsx', column_names)  # Берем первый файл из однотипных для первой строки
    result_sheet = get_first_sheet(result)
    total_files_load = 0
    for xlsx_doc in xlsx_docs:
        try:
            wb = load_workbook(xlsx_doc)  # Это все в цикле а имена файлов это массив для обработки
        except xml.etree.ElementTree.ParseError:
            print(f"Ошибка при чтении xlsx файла {xlsx_doc}")
            continue
        sheet = get_first_sheet(wb)
        column_indexes = get_sheet_indexes(sheet, IGNORE_LIST)
        print(column_indexes)
        for row in sheet.iter_rows(min_row=2, values_only=True):
            row_data = []
            for index in column_indexes:
                row_data.append(row[index - 1])
            result_sheet.append(row_data)
        result.save('Result.xlsx')
        total_files_load += 1
        print(f'Загружен {xlsx_doc}, {total_files_load} из {len(xlsx_docs)}')

# как дополнять xlsx файл и сохранять его, затем опять
# как то избежать сохранение tags, возможно просто удалив его
